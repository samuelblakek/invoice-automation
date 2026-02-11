"""
Excel writer for updating PO records while preserving formatting.
"""

import logging
from pathlib import Path
from datetime import datetime
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill
from shutil import copy2

logger = logging.getLogger(__name__)

# Light blue fill applied to rows updated by automation
_UPDATED_ROW_FILL = PatternFill(
    start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
)


class ExcelWriter:
    """Writer for updating Excel workbooks while preserving formatting."""

    def __init__(self, workbook_path: Path, create_backup: bool = True):
        """
        Initialize Excel writer.

        Args:
            workbook_path: Path to the Excel workbook
            create_backup: Whether to create a backup before modifying
        """
        self.workbook_path = Path(workbook_path)
        self.workbook = None
        self.modified = False

        # Create backup if requested
        if create_backup:
            backup_path = self.workbook_path.with_suffix(".backup.xlsx")
            copy2(self.workbook_path, backup_path)
            logger.info("Created backup: %s", backup_path)

        # Load workbook
        self.workbook = openpyxl.load_workbook(self.workbook_path)

    def update_po_record(
        self,
        sheet_name: str,
        row_index: int,
        invoice_number: str,
        invoice_amount: Decimal,
        invoice_signed_date: datetime,
        nominal_code: str = "",
    ) -> bool:
        """
        Update a PO record with invoice details.

        Args:
            sheet_name: Name of the sheet
            row_index: Row index (0-based from pandas)
            invoice_number: Invoice number
            invoice_amount: Invoice amount (ex-VAT)
            invoice_signed_date: Date invoice was signed
            nominal_code: Nominal code to write (only if cell is empty)

        Returns:
            True if update successful, False otherwise
        """
        try:
            # Get the worksheet
            if sheet_name not in self.workbook.sheetnames:
                logger.error("Sheet '%s' not found", sheet_name)
                return False

            ws = self.workbook[sheet_name]

            # Convert pandas row index to Excel row number
            # Pandas row index 0 = Excel row 2 (assuming row 1 is header)
            # However, we need to account for possible title rows
            # Let's search for the header row first
            header_row = self._find_header_row(ws)
            if header_row is None:
                logger.error("Could not find header row in sheet '%s'", sheet_name)
                return False

            # Actual Excel row = header_row + row_index + 1
            excel_row = header_row + row_index + 1

            # Find column indices for the fields we need to update
            col_invoice_no = self._find_column(ws, header_row, "INVOICE NO.")
            col_invoice_amount = self._find_column(
                ws, header_row, "INVOICE AMOUNT (EX VAT)"
            )
            col_invoice_signed = self._find_column(ws, header_row, "INVOICE SIGNED")

            if not all([col_invoice_no, col_invoice_amount, col_invoice_signed]):
                logger.error(
                    "Could not find required columns in sheet '%s'", sheet_name
                )
                return False

            # Update the cells
            ws.cell(row=excel_row, column=col_invoice_no).value = invoice_number
            ws.cell(row=excel_row, column=col_invoice_amount).value = float(
                invoice_amount
            )
            ws.cell(
                row=excel_row, column=col_invoice_signed
            ).value = invoice_signed_date

            # Write nominal code if provided and cell is currently empty
            if nominal_code:
                col_nominal = self._find_column(ws, header_row, "NOMINAL CODE")
                if col_nominal:
                    existing = ws.cell(row=excel_row, column=col_nominal).value
                    if not existing or str(existing).strip() == "":
                        ws.cell(row=excel_row, column=col_nominal).value = nominal_code

            # Highlight the entire row light blue to mark it as processed
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = _UPDATED_ROW_FILL

            self.modified = True
            logger.info("Updated row %d in sheet '%s'", excel_row, sheet_name)
            return True

        except Exception:
            logger.exception("Error updating PO record")
            return False

    def _find_header_row(self, ws) -> Optional[int]:
        """
        Find the header row in a worksheet.

        Searches for a row containing 'PO' or 'INVOICE NO.' as indicators.

        Args:
            ws: Worksheet object

        Returns:
            Row number (1-based), or None if not found
        """
        for row_num in range(1, min(20, ws.max_row + 1)):  # Check first 20 rows
            for cell in ws[row_num]:
                if cell.value and isinstance(cell.value, str):
                    value_upper = cell.value.strip().upper()
                    # Require exact match for 'PO' to avoid matching titles
                    # like "Maintenance - PO's & Outstanding..."
                    if value_upper == "PO" or "INVOICE NO" in value_upper:
                        return row_num
        return None

    def _find_column(self, ws, header_row: int, column_name: str) -> Optional[int]:
        """
        Find the column index for a given column name.

        Args:
            ws: Worksheet object
            header_row: Row number containing headers (1-based)
            column_name: Name of the column to find

        Returns:
            Column number (1-based), or None if not found
        """
        column_name_upper = column_name.upper()
        for cell in ws[header_row]:
            if cell.value and isinstance(cell.value, str):
                if column_name_upper in cell.value.upper():
                    return cell.column
        return None

    def save(self) -> bool:
        """
        Save the workbook.

        Returns:
            True if save successful, False otherwise
        """
        if not self.modified:
            logger.debug("No changes to save")
            return True

        try:
            self.workbook.save(self.workbook_path)
            logger.info("Saved changes to %s", self.workbook_path)
            return True
        except Exception:
            logger.exception("Error saving workbook")
            return False

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        if exc_type is None and self.modified:
            self.save()
        self.close()
